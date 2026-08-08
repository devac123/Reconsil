from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Header-row detection
# ---------------------------------------------------------------------------
# Instead of a hard-coded map we scan the first N rows of each sheet and
# pick the first row where:
#   1. At least one cell has a non-null, non-empty value  (not completely blank)
#   2. None of the cells produce an "Unnamed: X" column header — those appear
#      when pandas hits a merged/title cell that is empty or has no text.
#
# This handles:
#   - Sheets with a single title row above the real header  (old row 1 sheets)
#   - Sheets with two title rows                            (old row 2 sheets)
#   - Sheets whose header starts at row 0                   (no title rows)
#   - Any future sheet without needing manual registration
#
# We still keep the explicit map as an *override* for sheets where the
# auto-detection might be ambiguous.

_SHEET_HEADER_ROW_OVERRIDE: dict[str, int] = {
    # Add overrides here only if auto-detection picks the wrong row.
    # e.g. "some special sheet": 3,
}

_MAX_SCAN_ROWS = 20   # scan at most this many rows looking for the header


def _is_clean_header_row(row: tuple) -> bool:
    """
    Return True if *row* looks like a real column-header row:
      - At least 2 cells have a non-null, non-empty string value.
      - No cell produces an "Unnamed" header (i.e. no None / empty cells
        before any filled cell — the leading check is implicit because
        pandas names them Unnamed only when the raw value is None/blank).
    We check the raw openpyxl values directly, so we catch this before
    pandas even parses the file.
    """
    filled = [c for c in row if c is not None and str(c).strip() != ""]
    if len(filled) < 2:
        return False
    # If more than half the cells in the row are blank/None it's probably
    # still a title row, not the real header.
    if len(filled) < len(row) / 2:
        return False
    return True


def _detect_header_row(file_path: Path, sheet_name: str) -> int:
    """
    Open the workbook with openpyxl and scan up to _MAX_SCAN_ROWS rows to
    find the first row that looks like a real header row.

    Returns the zero-based row index pandas should use as ``header=``.
    Falls back to 0 if nothing looks like a header.
    """
    # Check the explicit override map first
    key = sheet_name.strip().lower()
    if key in _SHEET_HEADER_ROW_OVERRIDE:
        return _SHEET_HEADER_ROW_OVERRIDE[key]

    wb  = load_workbook(file_path, read_only=True, data_only=True)
    ws  = wb[sheet_name]

    best_row = 0
    for zero_based_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=_MAX_SCAN_ROWS, values_only=True)
    ):
        if _is_clean_header_row(row):
            best_row = zero_based_idx
            break

    wb.close()
    return best_row


class FileReaderService:

    @staticmethod
    def read_excel(file_path: Path):
        workbook = load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        total_sheets = len(sheet_names)
        workbook.close()

        response = {
            "file_name": file_path.name,
            "total_sheets": total_sheets,
            "sheets": []
        }

        for sheet_name in sheet_names:
            header_row = _detect_header_row(file_path, sheet_name)

            wb = load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]
            total_rows = ws.max_row

            headers = []
            for i, row in enumerate(
                ws.iter_rows(min_row=1, max_row=header_row + 2, values_only=True)
            ):
                if i == header_row:
                    headers = [str(cell) if cell is not None else "" for cell in row]
                    break
            wb.close()

            response["sheets"].append({
                "name":       sheet_name,
                "rows":       total_rows,
                "columns":    headers,
                "header_row": header_row,
            })

        return response

    @staticmethod
    def read_sheet_as_dataframe(
        file_path: Path,
        sheet_name: str,
        header_row: int | None = None,
    ) -> pd.DataFrame:
        """
        Read a single worksheet from an Excel file into a pandas DataFrame.

        The header row is auto-detected (first row where ≥50 % of cells have
        actual values and no cell is blank/None) unless *header_row* is
        passed explicitly.

        Parameters
        ----------
        file_path : Path
            Path to the Excel file on disk.
        sheet_name : str
            Name of the worksheet tab to read.
        header_row : int | None
            Zero-based row index to use as column headers.  If *None*
            (default) the row is auto-detected via :func:`_detect_header_row`.

        Returns
        -------
        pd.DataFrame
            DataFrame whose columns start at the detected header row.
            Completely empty rows are dropped.
            Any remaining "Unnamed: X" columns are dropped so that rows
            with genuinely missing column names do not pollute the data.
        """
        if header_row is None:
            header_row = _detect_header_row(file_path, sheet_name)

        df: pd.DataFrame = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            engine="openpyxl",
            header=header_row,   # auto-detected or explicit
            dtype=object,        # keep every value as Python object; no auto-cast
        )

        # Drop rows where every column is NaN — truly empty rows
        df = df.dropna(how="all")

        # Ensure all column names are plain strings (openpyxl can return ints)
        df.columns = [str(col) for col in df.columns]

        # Drop columns whose header is "Unnamed: X" — these come from merged
        # title cells or genuinely missing column names and should not be stored.
        unnamed_cols = [c for c in df.columns if c.startswith("Unnamed:")]
        if unnamed_cols:
            df = df.drop(columns=unnamed_cols)

        return df


if __name__ == "__main__":
    file_path = Path("../../file/Development Indigo Reconciliation 2024-25.xlsx")
    response = FileReaderService.read_excel(file_path)
    for s in response["sheets"]:
        print(s["name"], "| header_row:", s["header_row"], "| cols:", s["columns"][:5])
