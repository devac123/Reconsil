"""
Reconciliation Routes
---------------------
Endpoints that trigger reconciliation and serve the result download.

POST /files/{uploaded_file_id}/reconcile
    Run the reconciliation engine for an uploaded file.
    Returns a JSON summary with row counts.

GET  /files/{uploaded_file_id}/reconcile/download
    Stream an Excel workbook containing the reconciliation result sheet.
    Column layout mirrors the original "Reconcilation" sheet in the
    client workbook.
"""

import io
import logging
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Path, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database.session import get_db
from app.models.reconciliation_result import ReconciliationResult
from app.models.reconciliation_remark import ReconciliationRemark
from app.models.uploaded_file import UploadedFile
from app.service.reconciliation_service import ReconciliationService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/files", tags=["Reconciliation"])


class CombinedReconciliationRequest(BaseModel):
    uploaded_file_ids: list[int] = Field(..., min_length=1)
    result_uploaded_file_id: int | None = None


# ─────────────────────────────────────────────────────────────────────────────
# Helper
# ─────────────────────────────────────────────────────────────────────────────

def _get_file_or_404(uploaded_file_id: int, db: Session) -> UploadedFile:
    f = db.query(UploadedFile).filter(UploadedFile.id == uploaded_file_id).first()
    if not f:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"UploadedFile id={uploaded_file_id} not found.",
        )
    return f


# ─────────────────────────────────────────────────────────────────────────────
# POST /files/{id}/reconcile
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/reconcile-combined",
    status_code=status.HTTP_200_OK,
    summary="Run reconciliation across multiple uploaded workbooks",
    description=(
        "Combines rows from matching sheet names across multiple uploaded files "
        "before running the Cost-vs-Revenue reconciliation. Results are stored "
        "against result_uploaded_file_id, or the first uploaded_file_id when not supplied."
    ),
)
def run_combined_reconciliation(
    body: CombinedReconciliationRequest,
    db: Session = Depends(get_db),
):
    result_file_id = body.result_uploaded_file_id or body.uploaded_file_ids[0]

    try:
        svc = ReconciliationService(db)
        total_rows = svc.reconcile_combined(
            uploaded_file_ids=body.uploaded_file_ids,
            result_uploaded_file_id=result_file_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc))
    except Exception as exc:
        logger.exception(
            "Combined reconciliation failed for uploaded_file_ids=%s.",
            body.uploaded_file_ids,
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Combined reconciliation failed: {exc}",
        )

    return {
        "uploaded_file_ids": body.uploaded_file_ids,
        "result_uploaded_file_id": result_file_id,
        "status": "completed",
        "reconciled_rows": total_rows,
        "message": (
            f"Combined reconciliation complete. {total_rows} PNR rows produced "
            f"from {len(set(body.uploaded_file_ids))} workbook(s)."
        ),
    }

@router.post(
    "/{uploaded_file_id}/reconcile",
    status_code=status.HTTP_200_OK,
    summary="Run reconciliation for an uploaded file",
    description=(
        "Triggers the Cost-vs-Revenue reconciliation engine for the given "
        "uploaded file. Any previously stored results for this file are "
        "replaced. Returns a summary with the number of reconciled PNR rows."
    ),
)
def run_reconciliation(
    uploaded_file_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
):
    _get_file_or_404(uploaded_file_id, db)

    try:
        svc = ReconciliationService(db)
        total_rows = svc.reconcile(uploaded_file_id)
    except Exception as exc:
        logger.exception(
            "Reconciliation failed for uploaded_file_id=%s.", uploaded_file_id
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Reconciliation failed: {exc}",
        )

    return {
        "uploaded_file_id": uploaded_file_id,
        "status":           "completed",
        "reconciled_rows":  total_rows,
        "message":          f"Reconciliation complete. {total_rows} PNR rows produced.",
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /files/{id}/reconcile/results  — filtered, paginated JSON query
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/{uploaded_file_id}/reconcile/results",
    status_code=status.HTTP_200_OK,
    summary="Query reconciliation results with filters",
    description=(
        "Returns a paginated list of reconciled PNR rows. "
        "Supports filtering by PNR, remark, variance range, and per-source "
        "existence filters (cost_filter, cashx_filter, spyj_filter)."
    ),
)
def get_reconciliation_results(
    uploaded_file_id: int = Path(..., gt=0),
    # ── Filters ────────────────────────────────────────────────────────
    pnr: Optional[str] = Query(
        None,
        description="Filter by PNR (case-insensitive partial match).",
        examples=["KTCKMP"],
    ),
    remark: Optional[List[str]] = Query(
        None,
        description=(
            "Filter by one or more remarks (exact match, repeatable). "
            "e.g. ?remark=Matched&remark=Variance"
        ),
    ),
    variance_min: Optional[float] = Query(
        None,
        description="Minimum variance value (inclusive).",
    ),
    variance_max: Optional[float] = Query(
        None,
        description="Maximum variance value (inclusive).",
    ),
    # ── Comparison result filter ────────────────────────────────────────
    comparison: Optional[str] = Query(
        None,
        description=(
            "Filter by comparison outcome (all three sources present). "
            "'matched' = variance within tolerance; 'variance' = variance outside tolerance."
        ),
        pattern="^(matched|variance)$",
    ),
    # ── Source-existence filters ────────────────────────────────────────
    cost_filter: Optional[str] = Query(
        None,
        description=(
            "Filter by Cost source existence. "
            "'exist' = record found in Cost; 'not_exist' = not found in Cost."
        ),
        pattern="^(exist|not_exist)$",
    ),
    cashx_filter: Optional[str] = Query(
        None,
        description=(
            "Filter by Cash X source existence. "
            "'exist' = record found in Cash X; 'not_exist' = not found."
        ),
        pattern="^(exist|not_exist)$",
    ),
    spyj_filter: Optional[str] = Query(
        None,
        description=(
            "Filter by SPYJ source existence. "
            "'exist' = record found in SPYJ; 'not_exist' = not found."
        ),
        pattern="^(exist|not_exist)$",
    ),
    # ── Pagination ─────────────────────────────────────────────────────
    page: int = Query(1, ge=1, description="Page number (1-based)."),
    page_size: int = Query(50, ge=1, le=500, description="Rows per page (max 500)."),
    db: Session = Depends(get_db),
):
    _get_file_or_404(uploaded_file_id, db)

    q = db.query(ReconciliationResult).filter(
        ReconciliationResult.uploaded_file_id == uploaded_file_id
    )

    # ── Apply filters ──────────────────────────────────────────────────
    if pnr:
        q = q.filter(ReconciliationResult.pnr.ilike(f"%{pnr}%"))

    if remark:
        # Join to remarks table and match any of the requested labels
        q = (
            q.join(ReconciliationRemark, ReconciliationRemark.result_id == ReconciliationResult.id)
            .filter(ReconciliationRemark.remark.in_(remark))
            .distinct()
        )

    if variance_min is not None:
        q = q.filter(ReconciliationResult.variance >= variance_min)

    if variance_max is not None:
        q = q.filter(ReconciliationResult.variance <= variance_max)

    # ── Source-existence filters (AND logic, independent per source) ───
    # cost_pnr == "not found" means the PNR was absent from the Cost sheet
    if cost_filter == "exist":
        q = q.filter(ReconciliationResult.cost_pnr != "not found")
    elif cost_filter == "not_exist":
        q = q.filter(ReconciliationResult.cost_pnr == "not found")

    if cashx_filter == "exist":
        q = q.filter(ReconciliationResult.cashx_pnr != "not found")
    elif cashx_filter == "not_exist":
        q = q.filter(ReconciliationResult.cashx_pnr == "not found")

    if spyj_filter == "exist":
        q = q.filter(ReconciliationResult.spyj_pnr != "not found")
    elif spyj_filter == "not_exist":
        q = q.filter(ReconciliationResult.spyj_pnr == "not found")

    # ── Comparison filter (Matched / Variance) ─────────────────────────
    # Applies only to rows where all three sources are present.
    # "Matched" and "Variance" are the remark values assigned by the engine
    # when no source is missing.
    if comparison == "matched":
        q = q.filter(ReconciliationResult.remark == "Matched")
    elif comparison == "variance":
        q = q.filter(ReconciliationResult.remark == "Variance")

    total = q.count()

    # ── Pagination ─────────────────────────────────────────────────────
    offset  = (page - 1) * page_size
    records = q.order_by(ReconciliationResult.pnr).offset(offset).limit(page_size).all()

    return {
        "uploaded_file_id": uploaded_file_id,
        "total":            total,
        "page":             page,
        "page_size":        page_size,
        "total_pages":      max(1, -(-total // page_size)),  # ceiling division
        "filters": {
            "pnr":          pnr,
            "remark":       remark,
            "variance_min": variance_min,
            "variance_max": variance_max,
            "cost_filter":  cost_filter,
            "cashx_filter": cashx_filter,
            "spyj_filter":  spyj_filter,
            "comparison":   comparison,
        },
        "results": [
            {
                "id":             r.id,
                "pnr":            r.pnr,
                "booking_date":   r.booking_date.isoformat() if r.booking_date else None,
                "booking_id":     r.booking_id,
                "customer_name":  r.customer_name,
                "cost_pnr":       r.cost_pnr,
                "cost_sale":      r.cost_sale,
                "cost_refund":    r.cost_refund,
                "cost_net":       r.cost_net,
                "cashx_pnr":      r.cashx_pnr,
                "cashx_client_name": r.cashx_client_name,
                "cashx_client_code": r.cashx_client_code,
                "cashx_amount":   r.cashx_amount,
                "cashx_refund":   r.cashx_refund,
                "cashx_net":      r.cashx_net,
                "spyj_pnr":       r.spyj_pnr,
                "spyj_client_name": r.spyj_client_name,
                "spyj_client_code": r.spyj_client_code,
                "spyj_amount":    r.spyj_amount,
                "spyj_refund":    r.spyj_refund,
                "spyj_net":       r.spyj_net,
                "variance":       r.variance,
                "remark":         r.remark,
                "remarks":        [rm.remark for rm in r.remarks],
                "revised_remark": r.revised_remark,
                "final_remark":   r.final_remark,
            }
            for r in records
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /files/{id}/reconcile/summary  — counts per remark
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/{uploaded_file_id}/reconcile/summary",
    status_code=status.HTTP_200_OK,
    summary="Reconciliation summary — counts and totals per remark",
    description=(
        "Returns the total PNR count, variance totals, and a breakdown "
        "of how many rows fall under each remark category."
    ),
)
def get_reconciliation_summary(
    uploaded_file_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
):
    _get_file_or_404(uploaded_file_id, db)

    # Total rows + overall variance sum
    totals = db.query(
        func.count(ReconciliationResult.id).label("total_pnrs"),
        func.sum(ReconciliationResult.variance).label("total_variance"),
        func.sum(ReconciliationResult.cost_net).label("total_cost_net"),
        func.sum(ReconciliationResult.cashx_net).label("total_cashx_net"),
        func.sum(ReconciliationResult.spyj_net).label("total_spyj_net"),
    ).filter(
        ReconciliationResult.uploaded_file_id == uploaded_file_id
    ).one()

    # Count + variance sum per remark (from the normalised remarks table)
    remark_rows = (
        db.query(
            ReconciliationRemark.remark,
            func.count(ReconciliationResult.id.distinct()).label("count"),
            func.sum(ReconciliationResult.variance).label("variance_total"),
        )
        .join(ReconciliationRemark, ReconciliationRemark.result_id == ReconciliationResult.id)
        .filter(ReconciliationResult.uploaded_file_id == uploaded_file_id)
        .group_by(ReconciliationRemark.remark)
        .order_by(func.count(ReconciliationResult.id.distinct()).desc())
        .all()
    )

    return {
        "uploaded_file_id": uploaded_file_id,
        "total_pnrs":       totals.total_pnrs or 0,
        "total_variance":   round(totals.total_variance or 0, 2),
        "total_cost_net":   round(totals.total_cost_net or 0, 2),
        "total_cashx_net":  round(totals.total_cashx_net or 0, 2),
        "total_spyj_net":   round(totals.total_spyj_net or 0, 2),
        "by_remark": [
            {
                "remark":          row.remark or "Unassigned",
                "count":           row.count,
                "variance_total":  round(row.variance_total or 0, 2),
            }
            for row in remark_rows
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /files/{id}/reconcile/download
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/{uploaded_file_id}/reconcile/download",
    summary="Download reconciliation result as Excel",
    description=(
        "Generates and streams an .xlsx file containing the reconciliation "
        "result sheet. The column layout mirrors the original client workbook."
    ),
)
def download_reconciliation(
    uploaded_file_id: int = Path(..., gt=0),
    db: Session = Depends(get_db),
):
    uploaded_file = _get_file_or_404(uploaded_file_id, db)

    results: list[ReconciliationResult] = (
        db.query(ReconciliationResult)
        .filter(ReconciliationResult.uploaded_file_id == uploaded_file_id)
        .order_by(ReconciliationResult.pnr)
        .all()
    )

    if not results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=(
                "No reconciliation results found for this file. "
                "Please run reconciliation first (POST /files/{id}/reconcile)."
            ),
        )

    xlsx_bytes = _build_excel(results, uploaded_file)

    filename = (
        f"reconciliation_{uploaded_file.original_filename}"
        .replace(" ", "_")
        .replace(".xlsx", "")
        + "_result.xlsx"
    )

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

# ── colours (matching the style of the original sheet) ──
_HDR_FILL_1  = PatternFill("solid", fgColor="1F4E79")   # dark blue  – main groups
_HDR_FILL_2  = PatternFill("solid", fgColor="2E75B6")   # mid  blue  – sub-headers
_HDR_FILL_VAR = PatternFill("solid", fgColor="833C00")  # dark amber – variance
_WHITE_FONT  = Font(bold=True, color="FFFFFF")
_BOLD        = Font(bold=True)
_THIN = Side(border_style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_REMARK_COLOURS = {
    "Matched":                PatternFill("solid", fgColor="E2EFDA"),  # light green
    "Markup/Booking Charges": PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "Not in Cost":            PatternFill("solid", fgColor="FCE4D6"),  # light orange
    "Not in CASH X":          PatternFill("solid", fgColor="FCE4D6"),
    "Not in SPYJ":            PatternFill("solid", fgColor="FCE4D6"),
    "Variance":               PatternFill("solid", fgColor="FFDEDE"),  # light red
}


def _build_excel(
    results: list[ReconciliationResult],
    uploaded_file: UploadedFile,
) -> bytes:
    """Build and return the reconciliation Excel as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    # ── Title rows ──────────────────────────────────────────────────────
    ws.merge_cells("A1:V1")
    title_cell = ws["A1"]
    title_cell.value = "Indigo Reconciliation - Cost vs Revenue"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill  = _HDR_FILL_1
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:V2")
    sub_cell = ws["A2"]
    sub_cell.value = f"File: {uploaded_file.original_filename}"
    sub_cell.font  = Font(italic=True, color="595959")
    sub_cell.alignment = Alignment(horizontal="center")

    # ── Group header row (row 3) ─────────────────────────────────────────
    # Columns:
    #  A        = Parental PNR
    #  B        = Booking ID
    #  C–F      = Cost (AIR COST TRN)
    #  G–L      = CASH X (CASH x SAle / CASH X Re)
    #  M–R      = SPYJ Online Sale
    #  S        = Variance
    #  T–V      = Remarks

    group_headers = [
        (1,  1,  "Parental PNR",   _HDR_FILL_1),
        (2,  2,  "Booking ID",     _HDR_FILL_1),
        (3,  6,  "Cost",           _HDR_FILL_2),
        (7, 12,  "CASH X",         _HDR_FILL_2),
        (13, 18, "SPYJ Online Sale",_HDR_FILL_2),
        (19, 19, "VARIANCE",       _HDR_FILL_VAR),
        (20, 22, "Remarks",        _HDR_FILL_1),
    ]

    for start_col, end_col, label, fill in group_headers:
        if start_col == end_col:
            cell = ws.cell(row=3, column=start_col)
            cell.value = label
        else:
            ws.merge_cells(
                start_row=3, start_column=start_col,
                end_row=3,   end_column=end_col,
            )
            cell = ws.cell(row=3, column=start_col)
            cell.value = label

        cell.font      = _WHITE_FONT
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _BORDER

    ws.row_dimensions[3].height = 18

    # ── Column sub-headers (row 4) ───────────────────────────────────────
    col_headers = [
        "Parental PNR",               # A
        "Booking ID",                 # B
        "PNR",    "Sale",   "Refund",  "Net",     # C–F  Cost
        "PNR",    "Client Name", "Client Code", "Amount", "Refund",  "Net",     # G–L  CASH X
        "PNR",    "Client Name", "Client Code", "Amount", "Refund",  "Net",     # M–R  SPYJ
        "Cost−CashX−SPYJ",            # S  Variance
        "Remark", "Revised Remark", "Final Remark", # T–V
    ]

    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font      = _WHITE_FONT
        cell.fill      = _HDR_FILL_2
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER

    ws.row_dimensions[4].height = 30

    # ── Data rows (starting at row 5) ────────────────────────────────────
    for row_idx, rec in enumerate(results, start=5):
        remark_fill = _REMARK_COLOURS.get(rec.remark or "", None)

        row_data = [
            rec.pnr,       rec.booking_id,
            rec.cost_pnr,   rec.cost_sale,   rec.cost_refund,   rec.cost_net,
            rec.cashx_pnr,  rec.cashx_client_name, rec.cashx_client_code, rec.cashx_amount, rec.cashx_refund, rec.cashx_net,
            rec.spyj_pnr,   rec.spyj_client_name,  rec.spyj_client_code,  rec.spyj_amount,  rec.spyj_refund,  rec.spyj_net,
            rec.variance,
            rec.remark,     rec.revised_remark, rec.final_remark,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            # Colour the whole row by remark
            if remark_fill:
                cell.fill = remark_fill
            # Right-align numeric columns
            if col_idx in (4, 5, 6, 10, 11, 12, 16, 17, 18, 19):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, float):
                    cell.number_format = "#,##0.00"

    # ── Column widths ─────────────────────────────────────────────────────
    widths = {
        1: 14,  # PNR
        2: 16,                            # Booking ID
        3: 12,  4: 12,  5: 12,  6: 12,   # Cost
        7: 12,  8: 24,  9: 14,  10: 12,  11: 12, 12: 12,   # CASH X
        13: 12, 14: 24, 15: 14, 16: 12, 17: 12, 18: 12,  # SPYJ
        19: 16,                           # Variance
        20: 28, 21: 24, 22: 24,           # Remarks
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze panes below headers
    ws.freeze_panes = "B5"

    # ── Auto-filter on the data ───────────────────────────────────────────
    ws.auto_filter.ref = f"A4:{get_column_letter(22)}{4 + len(results)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
